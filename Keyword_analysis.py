import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from sys import argv
global wb
global ws
global files_name
script, file_dir = argv

def keywords_list(txt_name): #读取文件，并通过set()特性去重
    lines_seen = set()
    with open(txt_name, 'r') as f:
        data = f.readlines()  # txt中所有字符串读入data
        for line in data:
            lines_seen.add(line.strip("\n\r"))
        return lines_seen
        
def count(keywords, file_name, count_num):
    keywordslist = keywords_list(file_name)
    if keywords in keywordslist:
        count_num += 1
    else:
        pass
    return count_num

def takeSecond(elem):
    return elem[1]

def title_ins(file_path, column_num, count): # 加入标题行，统计关键词数量
    global wb
    global ws
    count = count - 2
    file_name = os.path.basename(os.path.realpath(file_path)) + '(' + str(count) + ')'
    tf = Font(size = 15)
    ws.cell(row = 1, column = column_num, value=file_name)
    ws.cell(row = 1, column = column_num).font = tf

def contrast(keyword, count_num): # 关键词在各文件出现次数统计
    global files_name
    for file_name in files_name:
        count_num = count(keyword, file_name, count_num)
    return count_num

def keyword_ins(): #排序并写入数据
    j = 1
    for file_name in files_name:
        Data_list = keywords_list(file_name)
        i = 2
        list_key = []
        for keyword in Data_list:
            count_num = 0
            count_num = contrast(keyword, count_num)
            list_key.append([keyword, count_num])
        list_key.sort(key=takeSecond, reverse=True)
        for keyword_list in list_key:
            num = keyword_list[1]
            keyword = keyword_list[0]
            keyword_frequency(num, keyword, i, j)
            i += 1
        title_ins(file_name, j, i)
        j += 1
    note(j)
    
def keyword_frequency(num, keyword, i, j): # 根据关键词出现次数，分配颜色
    global wb
    global ws
    ft = Font(color = 'FF5809')
    ft2 = Font(color = '46A3FF')
    ft3 = Font(color = '949449')
    if num >= 4:
        ws.cell(row = i, column = j, value = keyword)
        ws.cell(row = i, column = j).font = ft
    elif num == 3:
        ws.cell(row = i, column = j, value = keyword)
        ws.cell(row = i, column = j).font = ft2
    elif num == 2:
        ws.cell(row = i, column = j, value = keyword)
        ws.cell(row = i, column = j).font = ft3
    else:
        ws.cell(row = i, column = j, value = keyword) 

def note(x): #添加颜色说明
    ft = Font(color = 'FF5809')
    ft2 = Font(color = '46A3FF')
    ft3 = Font(color = '949449')
    tf = Font(size = 15)
    note1 = '关键词频次4次及以上'
    note2 = '关键词频次3次'
    note3 = '关键词频次2次'
    note4 = '可能的行业专属'
    ws.cell(row = 2, column = x, value = note1)
    ws.cell(row = 2, column = x).font = ft
    ws.cell(row = 3, column = x, value = note2)
    ws.cell(row = 3, column = x).font = ft2
    ws.cell(row = 4, column = x, value = note3)
    ws.cell(row = 4, column = x).font = ft3
    ws.cell(row = 5, column = x, value = note4)

def file_name_list(file_dir):
    files_name = []  
    for root, dirs, files in os.walk(file_dir):  
        for filename in files:
            if filename.endswith('txt'):
                files_name.append(os.path.realpath(file_dir) + '\\' + filename)
                print (filename)
            else:
	            print ('格式不对')
    return files_name
                   
def main():
    global wb
    global ws
    global files_name
    #file_dir = input("请输入需要处理文件夹位置:")
    wb = Workbook() #新建一个工作 
    #wb = openpyxl.load_workbook('result.xlsx')
    #ws = wb.active
    ws = wb.create_sheet(title = '数据结果')
    files_name = file_name_list(file_dir)
    keyword_ins()
    wb.save(file_dir + '//' + '数据分析结果.xlsx')
    print ('完成')

if __name__ == '__main__':
    main()
    input("按任意键继续")